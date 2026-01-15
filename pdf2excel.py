# pdf2excel.py

import pdfplumber
import pandas as pd
import os
import re
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time
from unidecode import unidecode

from quebec_regions_mapping import get_shore_region, get_custom_sector

def setup_logging():
    logs_dir = 'logs'
    os.makedirs(logs_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    log_file = os.path.join(logs_dir, f'conversion_log_{timestamp}.txt')
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    return log_file

def has_cpp_in_centris_no(centris_no_text):
    """Check if Centris No. cell contains CPP keyword (case-insensitive)."""
    if not centris_no_text:
        return False
    # Convert to string and check for CPP (case-insensitive)
    # CPP may appear with newlines, so we check the entire cell content
    text = str(centris_no_text).upper()
    return 'CPP' in text

def extract_with_pdfplumber(pdf_path):
    """Extracts rows from a PDF with columns [st, centris_no, municipality_borough, address, postal_code]."""
    with pdfplumber.open(pdf_path) as pdf:
        all_data = []
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                # table[0] might be a header row depending on the PDF format
                # If so, we skip it with table[1:], but adapt as needed
                for row in table[1:]:  # Skip the header row
                    # Join any split cells and clean up whitespace
                    cleaned_row = [' '.join(str(cell).split()) if cell else '' for cell in row]
                    # Filter out None/empty cells but keep the structure
                    if len(cleaned_row) >= 5:
                        # Take first 5 columns: ST, Centris No., Municipality/Borough, Address, Postal Code
                        all_data.append(cleaned_row[:5])
                    elif len(cleaned_row) == 4:
                        # Handle case where ST column might be missing (backward compatibility)
                        # Insert empty ST at the beginning
                        all_data.append([''] + cleaned_row[:4])
                    else:
                        logging.warning(f"Skipping malformed row: {cleaned_row}")
        return pd.DataFrame(all_data, columns=['st', 'centris_no', 'municipality_borough', 'address', 'postal_code'])

def extract_apartment(address):
    """Extract apartment substring (e.g. 'Apt. 101') from an address. Returns (address_without_apt, apartment_text)."""
    if not address:
        return ("", None)
    
    apt_index = address.lower().find('apt.')
    if apt_index == -1:
        return address, None
        
    base_address = address[:apt_index].rstrip(' ,')
    
    postal_pattern = r'[A-Z][0-9][A-Z]\s*[0-9][A-Z][0-9]'
    postal_match = re.search(postal_pattern, address[apt_index:])
    
    if postal_match:
        # Extract everything from 'apt.' up to the postal code
        apartment = address[apt_index:apt_index + postal_match.start()].strip()
        return base_address, apartment
    else:
        # If no postal code found, take the rest
        apartment = address[apt_index:].strip()
        return base_address, apartment

def clean_text(text, extract_apt=False, remove_accents=False):
    """Clean text, optionally remove accents, optionally handle apt extraction."""
    if not text:
        return ("", None) if extract_apt else ""
    
    # Attempt to fix partial prefix typos (example usage scenario)
    common_prefixes = {
        'ue ': 'Rue ',
        'v. ': 'Av. ',
        'h. ': 'Ch. ',
        'te ': 'Côte ',
        'l. ': 'Boul. '
    }
    for wrong, correct in common_prefixes.items():
        if text.lower().startswith(wrong):
            text = correct + text[len(wrong):]
            break
    
    # Remove text in parentheses (if needed)
    cleaned = re.sub(r'\s*\(.*$', '', text)

    if extract_apt:
        cleaned, apartment = extract_apartment(cleaned)
        # Clean trailing punctuation
        cleaned = re.sub(r'[\s\-,]+(?<!E)(?<!O)\.?$', '', cleaned).strip()
        if remove_accents:
            cleaned = unidecode(cleaned)
            if apartment:
                apartment = unidecode(apartment)
        return cleaned, apartment
    
    cleaned = re.sub(r'[\s\-,]+(?<!E)(?<!O)\.?$', '', cleaned).strip()
    if remove_accents:
        cleaned = unidecode(cleaned)
    return cleaned

def add_name_columns_to_df(df, merge_names, merged_name, column_names, default_values, remove_accents):
    """
    Adds name columns (merged or separate) to the DataFrame `df`,
    ensuring they're placed at the front of the DataFrame.
    """
    if df is None or df.empty:
        return df
    
    df = df.copy()
    
    if merge_names:
        # Merge First Name + Last Name into single column
        if 'First Name' in df.columns and 'Last Name' in df.columns:
            merged_col = (df['First Name'].fillna('') + ' ' + df['Last Name'].fillna('')).str.strip()
            df[merged_name] = merged_col
        else:
            # If first/last name columns are missing, fallback to default
            default_full_name = default_values.get(merged_name, "À l'occupant")
            df[merged_name] = [default_full_name]*len(df)
        
        if remove_accents:
            df[merged_name] = df[merged_name].astype(str).apply(lambda x: unidecode(x))
        
        # Reorder columns so the merged_name is at the front
        cols = df.columns.tolist()
        if merged_name in cols:
            cols.remove(merged_name)
            cols.insert(0, merged_name)
        df = df[cols]
    
    else:
        fn_key = 'First Name'
        ln_key = 'Last Name'
        col_first = column_names.get(fn_key, fn_key)
        col_last = column_names.get(ln_key, ln_key)

        if fn_key not in df.columns:
            df[col_first] = [default_values.get(col_first, "À l'occupant")] * len(df)
        else:
            df.rename(columns={fn_key: col_first}, inplace=True)
            df[col_first] = df[col_first].fillna(default_values.get(col_first, "À l'occupant"))
        
        if ln_key not in df.columns:
            df[col_last] = [default_values.get(col_last, "")] * len(df)
        else:
            df.rename(columns={ln_key: col_last}, inplace=True)
            df[col_last] = df[col_last].fillna(default_values.get(col_last, ""))

        if remove_accents:
            df[col_first] = df[col_first].astype(str).apply(lambda x: unidecode(x))
            df[col_last] = df[col_last].astype(str).apply(lambda x: unidecode(x))
        
        cols = df.columns.tolist()
        # Move the first/last name columns to the front
        if col_first in cols:
            cols.remove(col_first)
            cols.insert(0, col_first)
        if col_last in cols:
            cols.remove(col_last)
            cols.insert(1, col_last)
        df = df[cols]
    
    return df

def process_pdfs(
    pdf_paths,
    merge=False,
    column_names=None,
    merge_names=False, 
    merged_name="Full Name",
    default_values=None,
    file_format='xlsx',
    output_dir=None,
    custom_filename=None,
    merge_address=False,
    merged_address_name="Complete Address",
    address_separator=", ",
    province_default="QC",
    should_extract_apartment=False, 
    apartment_column_name="Apartment",
    filter_apartments=False, 
    include_apartment_column=True, 
    include_phone=False, 
    phone_default="", 
    include_date=False, 
    date_value=None, 
    filter_by_region=False, 
    region_branch_ids=None, 
    use_custom_sectors=False, 
    remove_accents=False,
    enable_logging=False
):
    """
    Main logic that processes PDF(s) and returns:
      - a list of DataFrames (all_dfs)
      - the final output directory (confirmed path)
    """
    if output_dir is None:
        output_dir = os.getcwd()  # Default to current directory if none provided
    
    # Ensure output_dir is absolute and exists
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)
    
    if enable_logging:
        logging.info(f"Processing PDFs with output_dir: {output_dir}")
    
    if column_names is None:
        column_names = {
            'First Name': 'First Name',
            'Last Name': 'Last Name',
            'Address': 'Address',
            'City': 'City',
            'Province': 'Province',
            'Postal Code': 'Postal Code'
        }
    if default_values is None:
        default_values = {}

    all_dfs = []

    for pdf_path in pdf_paths:
        logging.info(f"Processing PDF: {pdf_path}")
        df = extract_with_pdfplumber(pdf_path)
        logging.info(f"Extracted {len(df)} rows from {pdf_path}")

        # Filter based on ST (Status) column and CPP detection
        if 'st' in df.columns:
            # Clean ST column: strip whitespace and convert to uppercase for comparison
            df['st'] = df['st'].astype(str).str.strip().str.upper()
            
            # Filter: Include all 'SO' (Sold) rows, and 'AC' (Active) rows only if they have CPP
            # Create mask for rows to keep
            so_mask = df['st'] == 'SO'
            ac_with_cpp_mask = (df['st'] == 'AC') & df['centris_no'].apply(has_cpp_in_centris_no)
            
            # Apply filter
            df = df[so_mask | ac_with_cpp_mask].copy()
            # Reset index to avoid issues with iloc when addresses are unmerged
            df = df.reset_index(drop=True)
            
            logging.info(f"After ST/CPP filtering: {len(df)} rows remaining")
        else:
            logging.warning("ST column not found in extracted data. Skipping status filtering.")

        # Basic cleaning of municipality / address columns
        df['municipality_borough'] = df['municipality_borough'].apply(lambda x: x.split('(')[0].strip() if x else x)
        df['address'] = df['address'].apply(lambda x: x.strip() if x else x)
        
        # If address is empty, sometimes the PDF merges them incorrectly
        # This is a basic fallback example (may not be needed in all PDFs)
        df['address'] = df.apply(
            lambda row: row['address'] if row['address'] and row['address'].strip()
            else f"{row['municipality_borough']} {row['address']}",
            axis=1
        )

        # Optional region filtering
        if filter_by_region or use_custom_sectors:
            filtered_df = pd.DataFrame(columns=df.columns)
            for _, row in df.iterrows():
                city = row['municipality_borough']
                if use_custom_sectors:
                    sector = get_custom_sector(city, row['postal_code'])
                    if sector and sector in region_branch_ids:
                        row_df = pd.DataFrame([row])
                        row_df['Branch ID'] = region_branch_ids[sector]
                        filtered_df = pd.concat([filtered_df, row_df], ignore_index=True)
                else:
                    region = get_shore_region(city)
                    branch_id = region_branch_ids.get(f'flyer_{region}', region_branch_ids.get('flyer_unknown', 'unknown'))
                    if branch_id != 'unknown':
                        row_df = pd.DataFrame([row])
                        row_df['Branch ID'] = branch_id
                        filtered_df = pd.concat([filtered_df, row_df], ignore_index=True)
            
            if len(filtered_df) > 0:
                df = filtered_df.copy()
            else:
                logging.error("No valid rows after region filtering.")
                all_dfs.append(pd.DataFrame())
                continue

        # Build final output DataFrame
        output_df_final = None

        # (A) If MERGE_ADDRESS is True
        if merge_address:
            merged_addresses = []
            apartments = []
            branch_ids = []
            valid_indices = []

            for idx, row in df.iterrows():
                if should_extract_apartment:
                    clean_addr, apt = extract_apartment(row['address'])
                    if filter_apartments and apt is not None:
                        logging.info(f"Filtering out address with apartment: {row['address']} (apt={apt})")
                        continue
                    address_parts = [
                        clean_addr.strip() if clean_addr else "",
                        row['municipality_borough'].strip() if row['municipality_borough'] else "",
                        province_default,
                        row['postal_code']
                    ]
                    merged_address = address_separator.join(filter(None, address_parts))
                    merged_address = merged_address.strip()
                    
                    # Avoid duplicates in the final list
                    if merged_address not in merged_addresses:
                        merged_addresses.append(merged_address)
                        if 'Branch ID' in row:
                            branch_ids.append(row['Branch ID'])
                        if include_apartment_column and not filter_apartments:
                            apartments.append(apt)
                        valid_indices.append(idx)
                else:
                    if filter_apartments:
                        _, apt = extract_apartment(row['address'])
                        if apt is not None:
                            logging.info(f"Filtering out address with apartment: {row['address']}")
                            continue
                    plain_addr = clean_text(row['address'], extract_apt=False, remove_accents=remove_accents)
                    address_parts = [
                        plain_addr,
                        row['municipality_borough'],
                        province_default,
                        row['postal_code']
                    ]
                    merged_address = address_separator.join(filter(None, address_parts)).strip()
                    
                    if merged_address not in merged_addresses:
                        merged_addresses.append(merged_address)
                        if 'Branch ID' in row:
                            branch_ids.append(row['Branch ID'])
                        valid_indices.append(idx)
            
            if valid_indices:
                df_filtered = df.loc[valid_indices].copy()
                base_length = len(df_filtered)
                
                output_data = {}
                if 'Branch ID' in df_filtered.columns:
                    output_data['Branch ID'] = df_filtered['Branch ID'].tolist()
                elif branch_ids:
                    output_data['Branch ID'] = branch_ids[:base_length]
                
                # Remove accents if needed
                cleaned_merged_addresses = []
                for addr in merged_addresses[:base_length]:
                    cleaned_merged_addresses.append(unidecode(addr) if remove_accents else addr)
                
                output_data[merged_address_name] = cleaned_merged_addresses

                partial_df = pd.DataFrame(output_data)
                output_df_final = partial_df
            else:
                output_df_final = pd.DataFrame()
        
        # (B) If MERGE_ADDRESS is False
        else:
            valid_indices = []
            cleaned_addresses = []
            apartments = []

            for idx, row in df.iterrows():
                if should_extract_apartment:
                    clean_addr, apt = clean_text(row['address'], extract_apt=True, remove_accents=remove_accents)
                    if filter_apartments and apt is not None:
                        logging.info(f"Filtering out address with apartment: {row['address']}")
                        continue
                    cleaned_addresses.append(clean_addr)
                    if include_apartment_column and not filter_apartments:
                        apartments.append(apt)
                    valid_indices.append(idx)
                else:
                    if filter_apartments:
                        _, apt = extract_apartment(row['address'])
                        if apt is not None:
                            logging.info(f"Filtering out address with apartment: {row['address']}")
                            continue
                    plain_addr = clean_text(row['address'], extract_apt=False, remove_accents=remove_accents)
                    cleaned_addresses.append(plain_addr)
                    valid_indices.append(idx)
            
            df_filtered = df.loc[valid_indices].copy()
            output_data = {}
            
            # Build columns
            output_data[column_names['Address']] = cleaned_addresses
            output_data[column_names['City']] = df_filtered['municipality_borough'].tolist()
            output_data[column_names['Province']] = [
                default_values.get(column_names['Province'], province_default)
            ] * len(df_filtered)
            output_data[column_names['Postal Code']] = df_filtered['postal_code'].tolist()
            
            # If Branch ID was added
            if 'Branch ID' in df_filtered.columns:
                output_data['Branch ID'] = df_filtered['Branch ID'].tolist()

            if include_apartment_column and not filter_apartments and should_extract_apartment:
                output_data[apartment_column_name] = apartments
            
            partial_df = pd.DataFrame(output_data)

            # If address starts with the city name, remove duplication:
            address_col = column_names['Address']
            city_col = column_names['City']
            if address_col in partial_df.columns and city_col in partial_df.columns:
                partial_df[address_col] = partial_df.apply(
                    lambda row: row[address_col].replace(row[city_col], '', 1).strip()
                    if row[address_col].startswith(row[city_col]) else row[address_col],
                    axis=1
                )
            
            output_df_final = partial_df
        
        # Add name columns if we have a valid DF
        if output_df_final is not None and not output_df_final.empty:
            output_df_final = add_name_columns_to_df(
                df=output_df_final,
                merge_names=merge_names,
                merged_name=merged_name,
                column_names=column_names,
                default_values=default_values,
                remove_accents=remove_accents
            )

            # Phone column
            if include_phone:
                phone_col = column_names.get('Phone', 'Phone')
                output_df_final[phone_col] = [phone_default]*len(output_df_final)
            
            # Date column
            if include_date:
                date_col = column_names.get('Date', 'Date')
                output_df_final[date_col] = [date_value]*len(output_df_final)

            # Sort final
            if filter_by_region and 'Branch ID' in output_df_final.columns:
                sort_columns = ['Branch ID']
                if merge_address and merged_address_name in output_df_final.columns:
                    sort_columns.append(merged_address_name)
                else:
                    if column_names['City'] in output_df_final.columns:
                        sort_columns.append(column_names['City'])
                    if column_names['Address'] in output_df_final.columns:
                        sort_columns.append(column_names['Address'])
                output_df_final = output_df_final.sort_values(sort_columns)
            else:
                sort_col = merged_address_name if merge_address else column_names.get('City')
                if sort_col and sort_col in output_df_final.columns:
                    output_df_final = output_df_final.sort_values(by=sort_col)
        else:
            output_df_final = pd.DataFrame()

        all_dfs.append(output_df_final)

    return all_dfs, output_dir

def auto_adjust_columns(filename, df=None):
    """Auto-adjust column widths for Excel or format CSV content if needed."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    
    if filename.endswith('.xlsx'):
        workbook = load_workbook(filename)
        worksheet = workbook.active

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(filename)
    
    elif filename.endswith('.csv') and df is not None:
        # For CSV, we can left-pad columns to align, but it's purely cosmetic
        formatted_df = df.copy()
        max_lengths = {}
        for c in formatted_df.columns:
            max_lengths[c] = max(
                formatted_df[c].astype(str).str.len().max(),
                len(str(c))
            )
        for c in formatted_df.columns:
            width = max_lengths[c]
            formatted_df[c] = formatted_df[c].astype(str).str.ljust(width)
        return formatted_df

def convert_pdf_to_excel(
    pdf_files,
    output_dir,
    merge_files=False,
    custom_filename=None,
    enable_logging=False,
    column_names=None,
    merge_names=False, 
    merged_name="Full Name",
    default_values=None,
    file_format='xlsx',
    merge_address=False,
    merged_address_name="Complete Address",
    address_separator=", ",
    province_default="QC",
    should_extract_apartment=False,
    apartment_column_name="Apartment",
    filter_apartments=False,
    include_apartment_column=True,
    include_phone=False,
    phone_default="",
    include_date=False,
    date_value=None,
    filter_by_region=False,
    region_branch_ids=None,
    use_custom_sectors=False,
    remove_accents=False
):
    """
    High-level function that calls process_pdfs() and then writes outputs.
    Yields progress (int) or the final filename (str).
    """
    if enable_logging:
        logging.info(f"Starting conversion with output_dir={output_dir}")
        logging.info(f"Absolute output_dir path: {os.path.abspath(output_dir)}")
    
    # Make sure the directory is ready
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)
    
    if enable_logging:
        logging.info(f"Verified output directory: {output_dir}")
    
    pdf_paths = [pdf_files] if isinstance(pdf_files, str) else pdf_files
    total_files = len(pdf_paths)
    
    all_unique_addresses = set()
    all_data = []
    
    for i, pdf_path in enumerate(pdf_paths):
        # Extract dataframes from each PDF
        dfs, confirmed_output_dir = process_pdfs(
            [pdf_path],
            column_names=column_names,
            merge_names=merge_names,
            merged_name=merged_name,
            default_values=default_values,
            file_format=file_format,
            output_dir=output_dir,  # we pass the user-chosen directory here
            custom_filename=custom_filename,
            merge_address=merge_address,
            merged_address_name=merged_address_name,
            address_separator=address_separator,
            province_default=province_default,
            should_extract_apartment=should_extract_apartment,
            apartment_column_name=apartment_column_name,
            filter_apartments=filter_apartments,
            include_apartment_column=include_apartment_column,
            include_phone=include_phone,
            phone_default=phone_default,
            include_date=include_date,
            date_value=date_value,
            filter_by_region=filter_by_region,
            region_branch_ids=region_branch_ids,
            use_custom_sectors=use_custom_sectors,
            remove_accents=remove_accents,
            enable_logging=enable_logging
        )
        
        # Either we are merging all into a single final file or separate outputs
        if merge_files:
            for df in dfs:
                if merged_address_name in df.columns:
                    # Make sure we only keep unique addresses if merging
                    unique_mask = ~df[merged_address_name].isin(all_unique_addresses)
                    all_unique_addresses.update(df[merged_address_name][unique_mask])
                    df_unique = df[unique_mask].copy()
                    if not df_unique.empty:
                        all_data.append(df_unique)
                else:
                    all_data.append(df)
        else:
            # If not merging, we just store each PDF's data in all_data
            all_data.extend(dfs)
        
        # Emit progress up to ~90% across the loop
        progress = int((i + 1) / total_files * 90)
        yield progress

    # After processing all PDFs, either write a single merged file or multiple files
    current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    final_filename = None

    if merge_files:
        if all_data:
            merged_df = pd.concat(all_data, ignore_index=True)
            
            # Format date column if it exists
            if include_date and 'Date' in merged_df.columns:
                merged_df['Date'] = pd.to_datetime(merged_df['Date']).dt.strftime('%Y-%m-%d')
            
            # Drop apartment column if it was only used internally
            if should_extract_apartment and not include_apartment_column and apartment_column_name in merged_df.columns:
                merged_df.drop(columns=[apartment_column_name], inplace=True, errors='ignore')

            # If filtering by region, sort by region + city/address
            if filter_by_region and 'Branch ID' in merged_df.columns:
                sort_cols = ['Branch ID']
                if merge_address and merged_address_name in merged_df.columns:
                    sort_cols.append(merged_address_name)
                else:
                    city_col = column_names.get('City')
                    addr_col = column_names.get('Address')
                    if city_col and city_col in merged_df.columns:
                        sort_cols.append(city_col)
                    if addr_col and addr_col in merged_df.columns:
                        sort_cols.append(addr_col)
                merged_df = merged_df.sort_values(sort_cols)
            else:
                # Otherwise just sort by city or merged_address
                sort_col = merged_address_name if merge_address else column_names.get('City')
                if sort_col and sort_col in merged_df.columns:
                    merged_df = merged_df.sort_values(by=sort_col)

            if custom_filename:
                output_filename = os.path.join(confirmed_output_dir, f'{custom_filename}.{file_format}')
            else:
                output_filename = os.path.join(confirmed_output_dir, f'merged_output_{current_time}.{file_format}')

            # Save
            if file_format == 'xlsx':
                merged_df.to_excel(output_filename, index=False)
                auto_adjust_columns(output_filename)
            else:
                formatted_df = auto_adjust_columns(output_filename, merged_df)
                formatted_df.to_csv(output_filename, index=False, encoding='utf-8-sig')

            final_filename = output_filename
            yield output_filename
    else:
        # Not merging => each DF to its own file
        last_file = None
        for i, df in enumerate(all_data):
            if include_date and 'Date' in df.columns:
                df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d')
            
            if custom_filename:
                # e.g. custom_name_1.xlsx, custom_name_2.xlsx, ...
                output_filename = os.path.join(confirmed_output_dir, f'{custom_filename}_{i+1}.{file_format}')
            else:
                # e.g. from PDF base name plus timestamp
                base_name = os.path.splitext(os.path.basename(pdf_paths[i]))[0]
                output_filename = os.path.join(confirmed_output_dir, f'{base_name}_{current_time}.{file_format}')

            if filter_by_region and 'Branch ID' in df.columns:
                if merge_address and merged_address_name in df.columns:
                    df = df.sort_values(['Branch ID', merged_address_name])
                else:
                    df = df.sort_values(['Branch ID', column_names['City'], column_names['Address']])
            else:
                sort_col = merged_address_name if merge_address else column_names['City']
                if sort_col in df.columns:
                    df = df.sort_values(sort_col)

            # Save
            if file_format == 'xlsx':
                df.to_excel(output_filename, index=False)
                auto_adjust_columns(output_filename)
            else:
                formatted_df = auto_adjust_columns(output_filename, df)
                formatted_df.to_csv(output_filename, index=False, encoding='utf-8-sig')
            
            last_file = output_filename
        
        final_filename = last_file
        yield last_file

    yield 100  # final progress

    if enable_logging:
        if final_filename:
            logging.info(f"Conversion complete. Final file: {final_filename}")
        else:
            logging.info("Conversion complete with no output file.")