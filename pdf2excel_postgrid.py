import pdfplumber
import pandas as pd
import re
import requests
import time
from concurrent.futures import ThreadPoolExecutor
from tkinter import Tk, filedialog
import os
import requests_cache
import json
from city_mappings import get_city_from_borough
from dotenv import load_dotenv
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime
import sys
from io import StringIO
import builtins
from retry import retry
import difflib
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Create a logs directory if it doesn't exist
if not os.path.exists('logs'):
    os.makedirs('logs')

# Set up logging
current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file = f'logs/postgrid_api_{current_time}.log'

# Create a custom logger
logger = logging.getLogger('postgrid_api')
logger.setLevel(logging.DEBUG)

# Create handlers
file_handler = RotatingFileHandler(log_file, maxBytes=10485760, backupCount=5)
console_handler = logging.StreamHandler()

# Create formatters and add it to handlers
log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
file_formatter = logging.Formatter(log_format)
console_formatter = logging.Formatter('%(levelname)s: %(message)s')
file_handler.setFormatter(file_formatter)
console_handler.setFormatter(console_formatter)

# Add handlers to the logger
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# Modify the custom_print function
def custom_print(*args, **kwargs):
    message = ' '.join(map(str, args))
    logger.info(message)

# Replace the built-in print function
builtins.print = custom_print

# At the start of your main execution
logger.info("Starting PostGrid address processing")

# Enable caching for requests
requests_cache.install_cache('postgrid_cache', backend='sqlite', expire_after=86400)  # Cache expires after 1 day

# Load environment variables
load_dotenv()

# PostGrid API Key
API_KEY = os.getenv('POSTGRID_API_KEY')

# Function to get file paths using a file dialog
def get_pdf_paths():
    root = Tk()
    root.withdraw()  # Hide the main Tkinter window
    file_paths = filedialog.askopenfilenames(title="Select PDF Files", filetypes=[("PDF Files", "*.pdf")])
    return list(file_paths)

# Function to extract data using pdfplumber
def extract_with_pdfplumber(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_data = []
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    # Skip the first row (header) for each page
                    all_data.extend(table[1:])
            if all_data:
                combined_df = pd.DataFrame(all_data)
                return combined_df
    except Exception as e:
        print(f"PDFPlumber failed: {e}")
    return None

# Function to parallelize PDF extraction
def parallel_pdf_extraction(paths):
    with ThreadPoolExecutor(max_workers=5) as executor:
        return list(executor.map(extract_with_pdfplumber, paths))

# Function to separate apartment number from address
def separate_apartment(address):
    if address is None:
        return None, None
    
    address = clean_address(address)
    
    # Updated regex to capture more apartment formats
    apt_regex = re.compile(r'(.*?)\s*(?:,\s*)?((?:app?t?\.?\s*|unit\s*|suite\s*|#\s*)#?\s*\d+[a-zA-Z]?)\s*$', re.IGNORECASE)
    match = apt_regex.search(address)
    
    if match:
        main_address = match.group(1).strip()
        apt_number = match.group(2)
        
        if apt_number:
            # Extract only the numeric part (and potential letter) from the apartment number
            apt_number = re.sub(r'^(app?t?\.?\s*|unit\s*|suite\s*|#\s*)#?\s*', '', apt_number, flags=re.IGNORECASE).strip()
            return main_address, apt_number
    
    return address, None

# Function to format address for PostGrid
def format_address_for_postgrid(address, apartment):
    if apartment:
        return f"{apartment}-{address}"
    return address

# Add a retry decorator to handle temporary API failures
@retry(tries=3, delay=1, backoff=2)
def postgrid_api_call(url, method='post', **kwargs):
    response = requests.request(method, url, **kwargs)
    response.raise_for_status()
    return response.json()

def find_best_suggestion(input_address, input_city, suggestions):
    input_apt, input_number, input_street = extract_address_components(input_address)
    
    best_match = None
    highest_score = -1

    for suggestion in suggestions:
        suggestion_address = suggestion.get('line1', '').lower()
        sugg_apt, sugg_number, sugg_street = extract_address_components(suggestion_address)
        suggestion_city = suggestion.get('city', '').lower()

        score = 0

        # Check if the suggestion's street number matches the input or is within a range
        if sugg_number == input_number:
            score += 1000
        elif '...' in sugg_number:
            range_start, range_end = map(int, re.findall(r'\d+', sugg_number))
            if range_start <= int(input_number) <= range_end:
                score += 800
        else:
            continue  # Skip if street number doesn't match and isn't in range

        # Street name similarity
        street_similarity = difflib.SequenceMatcher(None, input_street, sugg_street).ratio()
        score += street_similarity * 100

        # City match
        if input_city.lower() == suggestion_city:
            score += 100
        elif input_city.lower() in suggestion_city or suggestion_city in input_city.lower():
            score += 50

        # Apartment number handling
        if input_apt:
            if sugg_apt and input_apt == sugg_apt:
                score += 200
            elif '...' in suggestion_address:
                score += 100  # It's potentially correct, but we're not sure

        if score > highest_score:
            highest_score = score
            best_match = suggestion

    return best_match

def extract_address_components(address):
    parts = re.split(r'[-\s]+', address.lower())
    apt_number = None
    street_number = None
    street_name = []

    for i, part in enumerate(parts):
        if part.isdigit():
            if not street_number:
                street_number = part
            elif not apt_number:
                apt_number = street_number
                street_number = part
        elif re.match(r'\d+[a-z]?', part) and not apt_number:
            apt_number = part
        else:
            street_name.append(part)

    return apt_number, street_number, ' '.join(street_name)

def postgrid_suggest_address(address, city):
    url = "https://api.postgrid.com/v1/addver/suggestions"
    headers = {
        "x-api-key": API_KEY,
        "Content-Type": "application/json"
    }
    
    mapped_city = get_city_from_borough(city)
    logger.info(f"Mapped city from '{city}' to '{mapped_city}'")
    
    full_address = f"{address}, {mapped_city}, QC, Canada"
    payload = {
        "address": full_address,
        "country": "CA",
        "maxResults": 10
    }
    
    logger.debug(f"Suggestions API request for: {full_address}")
    
    try:
        data = postgrid_api_call(url, json=payload, headers=headers)
        logger.debug(f"Suggestions API response: {json.dumps(data, indent=2)}")
        
        if data.get("status") == "success":
            suggestions = data.get("data", [])
            if suggestions:
                best_match = find_best_suggestion(address, mapped_city, suggestions)
                if best_match:
                    logger.info(f"Best matching suggestion found for '{address}' in '{mapped_city}': {best_match}")
                    return best_match
                else:
                    logger.warning(f"No suitable suggestion found for: {full_address}")
                    logger.debug(f"Input address components: {extract_address_components(address)}")
                    logger.debug(f"All suggestions: {json.dumps(suggestions, indent=2)}")
            else:
                logger.warning(f"No suggestions found for: {full_address}")
        else:
            logger.error(f"Suggestions API Error: {json.dumps(data, indent=2)}")
    except Exception as e:
        logger.exception(f"PostGrid Suggestions API error: {e}")
    
    return {}

# Update the get_postal_code function
def get_postal_code(address, city):
    suggested = postgrid_suggest_address(address, city)
    postal_code = suggested.get("postalOrZip", "")
    confidence = "high" if postal_code else "low"
    logger.info(f"Suggested postal code for {address}, {city}: {postal_code} (Confidence: {confidence})")
    return postal_code, confidence

# Update the postgrid_validate_addresses_batch function
def postgrid_validate_addresses_batch(addresses):
    url = "https://api.postgrid.com/v1/addver/verifications/batch"
    headers = {
        "x-api-key": API_KEY,
        "Content-Type": "application/json"
    }
    
    # Get postal codes first and map cities
    for address in addresses:
        mapped_city = get_city_from_borough(address['city'])
        address['city'] = mapped_city
        postal_code, confidence = get_postal_code(address['line1'], mapped_city)
        address['postalOrZip'] = postal_code
        address['confidence'] = confidence
    
    payload = {"addresses": addresses}
    
    try:
        data = postgrid_api_call(url, json=payload, headers=headers)
        logger.debug(f"Batch validation API response: {json.dumps(data, indent=2)}")
        
        if data.get("status") == "success":
            results = data.get("data", {}).get("results", [])
            if isinstance(results, list):
                for i, result in enumerate(results):
                    verified_address = result.get("verifiedAddress", {})
                    original_address = addresses[i]
                    
                    if verified_address:
                        logger.info(f"Processing verified address: {verified_address}")
                        if not verified_address.get("postalOrZip"):
                            verified_address["postalOrZip"] = original_address['postalOrZip']
                            logger.info(f"Using pre-fetched postal code: {verified_address['postalOrZip']}")
                        verified_address["confidence"] = original_address['confidence']
                    else:
                        logger.warning(f"No verified address for: {original_address['line1']}, {original_address['city']}")
                        verified_address = {
                            "line1": original_address['line1'],
                            "city": original_address['city'],
                            "provinceOrState": original_address['provinceOrState'],
                            "country": original_address['country'],
                            "postalOrZip": original_address['postalOrZip'],
                            "confidence": "low"
                        }
                    
                    result["verifiedAddress"] = verified_address
                return results
            else:
                logger.error(f"Unexpected results format: {results}")
        else:
            logger.error(f"Batch API Error: {json.dumps(data, indent=2)}")
    except Exception as e:
        logger.exception(f"PostGrid Batch API error: {e}")
    
    return []

def postgrid_autocomplete_address(address):
    url = "https://api.postgrid.com/v1/addver/completions"
    headers = {
        "x-api-key": API_KEY,
        "Content-Type": "application/x-www-form-urlencoded"
    }
    
    payload = {
        "partialStreet": address.get('line1', ''),
        "city": address.get('city', ''),
        "provinceOrState": address.get('provinceOrState', ''),
        "country": address.get('country', '')
    }
    
    try:
        data = postgrid_api_call(url, method='post', data=payload, headers=headers)
        if data.get("status") == "success":
            completions = data.get("data", [])
            if completions:
                return completions[0].get("address", {})
        else:
            logger.error(f"Autocomplete API Error: {json.dumps(data, indent=2)}")
    except Exception as e:
        logger.exception(f"PostGrid Autocomplete API error: {e}")
    
    return {}

def clean_address(address):
    if not address:
        return ""
    # Remove trailing comma and any following whitespace
    address = re.sub(r',\s*$', '', address.strip())
    # Remove any other trailing punctuation
    address = re.sub(r'[.,;:]+$', '', address)
    # Remove 'Apt' or 'Apt.' if it's at the end of the address
    address = re.sub(r'\s+(?:Apt\.?|Apartment)?\s*$', '', address, flags=re.IGNORECASE)
    # Remove any leading or trailing whitespace
    return address.strip()

def get_postal_code_for_address(address, city):
    formatted_address = f"{address}, {city}, QC, Canada"
    payload = {
        "address": formatted_address
    }
    
    # Try suggestions first
    suggested = postgrid_suggest_address(payload)
    if suggested.get("postalOrZip"):
        return suggested.get("postalOrZip")
    
    # If suggestions don't work, try autocomplete
    autocompleted = postgrid_autocomplete_address(payload)
    if autocompleted.get("postalOrZip"):
        return autocompleted.get("postalOrZip")
    
    return ""  # Return empty string if no postal code found

# Function to clean 'None' values
def clean_none(value):
    return '' if pd.isna(value) or value == 'None' else str(value)

# Main execution
if __name__ == "__main__":
    if not API_KEY:
        print("Error: PostGrid API key not found. Please set POSTGRID_API_KEY in your .env file.")
        exit(1)
    
    pdf_paths = get_pdf_paths()
    if not pdf_paths:
        print("No PDF files selected. Exiting.")
        exit(1)

    pdf_dataframes = parallel_pdf_extraction(pdf_paths)

    for pdf_path, df in zip(pdf_paths, pdf_dataframes):
        if df is not None:
            df.dropna(how="all", inplace=True)

            expected_headers = ['centris_no', 'st', 'mun_bor', 'address', 'price', 'rent_price', 'pt', 'bt', 'rms', 'bdrm', 'bath_pr', 'f-s', 'p', 'g']
            if len(df.columns) >= len(expected_headers):
                df.columns = expected_headers + [f'column_{i}' for i in range(len(expected_headers), len(df.columns))]
            else:
                df.columns = expected_headers[:len(df.columns)]

            df = df.reset_index(drop=True)

            # Take only the 10 rows for testing
            df = df.head(10)

            mun_bor_column = next((col for col in df.columns if 'mun' in col or 'bor' in col), None)
            address_column = next((col for col in df.columns if 'address' in col), None)

            if mun_bor_column and address_column:
                df = df[[mun_bor_column, address_column]]
                df.columns = ['original_mun_bor', 'address']

                addresses_to_validate = []
                for _, row in df.iterrows():
                    main_address, apartment = separate_apartment(row['address'])
                    formatted_address = format_address_for_postgrid(main_address, apartment)
                    addresses_to_validate.append({
                        "line1": formatted_address,
                        "city": row['original_mun_bor'],
                        "provinceOrState": "QC",
                        "country": "CA"
                    })

                # Perform batch validation for this PDF file
                if addresses_to_validate:
                    batch_results = postgrid_validate_addresses_batch(addresses_to_validate)
                    
                    if batch_results:
                        # Process batch results
                        validated_addresses = []
                        for i, result in enumerate(batch_results):
                            verified_address = result.get("verifiedAddress", {})
                            original_address = addresses_to_validate[i]
                            
                            if verified_address:
                                postgrid_city = verified_address.get("city", "")
                                mapped_city = get_city_from_borough(postgrid_city)
                                
                                verified_line1 = verified_address.get('line1', '')
                                apt_match = re.match(r'^(\d+[a-zA-Z]?)-(.+)$', verified_line1)
                                if apt_match:
                                    apt_number, main_address = apt_match.groups()
                                    full_address = f"{main_address}, Apt. {apt_number}"
                                else:
                                    full_address = verified_line1
                                
                                validated_address = {
                                    "address": full_address,
                                    "city": mapped_city,
                                    "province": verified_address.get("provinceOrState", ""),
                                    "postal_code": verified_address.get("postalOrZip", ""),
                                    "original_mun_bor": original_address['city'],
                                    "original_address": original_address['line1'],
                                    "confidence": verified_address.get("confidence", "low")
                                }
                                logger.info(f"Validated address: {validated_address}")
                                validated_addresses.append(validated_address)
                            else:
                                validated_addresses.append({
                                    "address": original_address['line1'],
                                    "city": get_city_from_borough(original_address['city']),
                                    "province": "QC",
                                    "postal_code": original_address['postalOrZip'],
                                    "original_mun_bor": original_address['city'],
                                    "original_address": original_address['line1'],
                                    "confidence": "low"
                                })
                        
                        # Create output DataFrame
                        output_df = pd.DataFrame(validated_addresses)
                        output_df['fnam'] = 'À'
                        output_df['lnam'] = "l'occupant"
                        output_df['add1'] = output_df['address'].apply(clean_address)
                        output_df['city'] = output_df['city']
                        output_df['prov'] = output_df['province']
                        output_df['pc'] = output_df['postal_code']

                        # Select only the required columns
                        output_df = output_df[['fnam', 'lnam', 'add1', 'city', 'prov', 'pc', 'confidence']]

                        # Clean the output
                        for col in output_df.columns:
                            output_df[col] = output_df[col].apply(clean_none)

                        # Remove rows where all address-related fields are empty
                        output_df = output_df[~(output_df['add1'].isna() & output_df['city'].isna() & output_df['prov'].isna() & output_df['pc'].isna())]

                        # Export the final DataFrame to an Excel file with highlighting
                        output_filename = f'output_excel/{os.path.splitext(os.path.basename(pdf_path))[0]}_listings_postgrid.xlsx'
                        
                        # Create a new workbook and select the active sheet
                        wb = Workbook()
                        ws = wb.active

                        # Write the header
                        for col, header in enumerate(output_df.columns, start=1):
                            ws.cell(row=1, column=col, value=header)

                        # Write the data and apply conditional formatting
                        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                        for row, data in enumerate(output_df.values, start=2):
                            for col, value in enumerate(data, start=1):
                                cell = ws.cell(row=row, column=col, value=value)
                                if col == 7 and value == 'low':  # 'confidence' column
                                    for data_cell in ws[row]:
                                        data_cell.fill = red_fill

                        # Save the workbook
                        wb.save(output_filename)

                        logger.info(f"Final output DataFrame:\n{output_df.to_string()}")

                        print(f"Excel file '{output_filename}' has been created successfully.")
                        print(f"Validated addresses: {output_df['add1'].notna().sum()}")
                        print(f"Total rows in output: {len(output_df)}")
                        print(f"Low confidence rows: {(output_df['confidence'] == 'low').sum()}")
                    else:
                        print(f"No addresses validated in {os.path.basename(pdf_path)}.")
                else:
                    print(f"No addresses to validate in {os.path.basename(pdf_path)}.")
            else:
                print(f"Required columns 'Mun/Bor.' and 'Address' not found in the extracted data for {os.path.basename(pdf_path)}.")
        else:
            print(f"No data extracted from the provided PDF file: {os.path.basename(pdf_path)}.")

    # At the end of your main execution
    logger.info("PostGrid address processing completed")
    print(f"Detailed API logs have been saved to: {log_file}")